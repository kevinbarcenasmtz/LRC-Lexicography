"""Build the DravidiLex pilot import files for the LRC platform.

Inputs
------
- data/dravidian/starling/output.xlsx
    Tree-structured Starling export (one row per word, linked by Parent Word ID),
    produced by src/dravidian/notebooks/destructuring_of_scraped_json.ipynb.
- data/dravidian/lrc_import/dravidilex_languages.csv
    Family / Subfamily / Language tiers for the LRC import. This tracked CSV is
    the cross-machine source of truth; the older
    data/dravidian/three-tier-language tree.xlsx is accepted as a fallback only.

Outputs (data/dravidian/lrc_import/)
------------------------------------
- dravidian_starling_data.xlsx
    Same tree as output.xlsx, but every protoform row carries the DED
    number(s) of all reflexes in its subtree (all distinct values kept).
- dravidilex_languages.csv
    Normalized Family,Subfamily,Language rows for the Laravel import command.
    Includes intermediate proto-languages that appear in the Starling data so
    protoform rows can resolve a Language at import time.
- dravidilex_batch_import.json
    Reflex rows in the Utilities uploader format (Headwords / Gloss /
    Language required; HeadwordEntries gives the Laravel importer already-split
    LexReflex.entries; every other key lands in LexReflexExtraData).
    The tree is preserved through Starling ID / Parent Word ID extra data.
- dravidilex_batch_import.xlsx
    Human-reviewable spreadsheet of the same rows.
"""

import csv
import html
import json
import re
from collections import defaultdict
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[3]
DATA_DIR = REPO_ROOT / "data" / "dravidian"
STARLING_XLSX = DATA_DIR / "starling" / "output.xlsx"
TREE_XLSX = DATA_DIR / "three-tier-language tree.xlsx"
OUT_DIR = DATA_DIR / "lrc_import"
LANGUAGES_CSV = OUT_DIR / "dravidilex_languages.csv"
# Scraped Burrow & Emeneau DEDR entries, keyed by DED number for the Sources
# column (data/dravidian/burrow_ded/burrow_corpus.cleaned.json).
BURROW_CORPUS = DATA_DIR / "burrow_ded" / "burrow_corpus.cleaned.json"

# Source abbreviations (must match the `code` column of the committed
# dravidilex_sources.csv that ImportDravidilexCSV seeds LexSource from).
SOURCE_DEDR = "DEDR"
SOURCE_CVOTGD = "CVOTGD"
SOURCE_STARLING = "STARLING"

# Roots/etyma cannot carry sources (no lex_etyma_source table), so their
# Starling provenance rides as a plain "Other Info" extra-data line instead.
ROOT_STARLING_ATTRIBUTION = "Starostin, Dravidian Etymology database (StarlingDB)"

SOURCE_HTML_ALLOWED_TAGS = {"a", "b", "br", "i", "small", "sup"}

STARLING_DATABASE_LABELS = {
    "/data/drav/dret": "Dravidian etymology",
    "/data/drav/sdret": "South Dravidian etymology",
    "/data/drav/gonet": "Gondwan etymology",
    "/data/drav/kogaet": "Kolami-Gadba etymology",
    "/data/drav/telet": "Telugu etymology",
    "/data/drav/kuiet": "Kui-Kuwi etymology",
    "/data/drav/ktet": "Kota-Toda etymology",
    "/data/drav/ndret": "North Dravidian etymology",
    "/data/drav/gndet": "Gondi etymology",
    "/data/drav/pemet": "Pengo-Manda etymology",
    "/data/drav/konet": "Konda etymology",
    "/data/drav/braet": "Brahui etymology",
}

# Starling uses both spellings; keep one.
LANGUAGE_NORMALIZATION = {
    "Proto-North-Dravidian": "Proto-North Dravidian",
}

# The scrape emitted a spurious *empty* placeholder for every North Dravidian
# etymon, tagged with the SPACED spelling "Proto-North Dravidian" — while the
# real etymon uses the HYPHENATED "Proto-North-Dravidian". The placeholder just
# echoes its parent root's gloss and carries no DED number, no source URL, no
# reflexes, and Depth 0. Because it shares an ID with the real (hyphenated) row,
# deduplicate_ids used to split the pair into e.g. PND008 / PND008-2, which is
# what surfaced as duplicate headwords on the site (two *aṭṭ-). All 563 spaced
# rows are empty (verified) and none have children, so we drop them at load time
# — before LANGUAGE_NORMALIZATION rewrites the hyphen and erases the only thing
# that tells artifact from real.
ARTIFACT_LANGUAGE = "Proto-North Dravidian"  # spaced spelling == the placeholder


def _is_blank(value):
    return value is None or str(value).strip() == ""


def is_nd_placeholder_artifact(row):
    """A row is a droppable North Dravidian placeholder iff it uses the spaced
    spelling AND carries no distinguishing data (no DED, no URL, Depth 0)."""
    return (
        row.get("Language") == ARTIFACT_LANGUAGE
        and _is_blank(row.get("Number in DED"))
        and _is_blank(row.get("URL"))
        and str(row.get("Depth")) == "0"
    )

# StarlingDB displays the alveolar series with a line below the letter; the
# scraper flattened that to "letter_" (e.g. ayyan̠ -> "ayyan_"). Map back to
# the DEDR line-below letters (precomposed where Unicode has them, combining
# U+0331 otherwise). Applied to every text column except IDs/URLs.
UNDERSCORE_LETTERS = {
    "r": "ṟ", "n": "ṉ", "d": "ḏ", "t": "ṯ", "l": "ḻ",
    "k": "ḵ", "h": "ẖ", "s": "s̱", "g": "g̱",
    "R": "Ṟ", "N": "Ṉ", "D": "Ḏ", "T": "Ṯ", "L": "Ḻ", "K": "Ḵ",
}
UNDERSCORE_RE = re.compile("([" + "".join(UNDERSCORE_LETTERS) + "])_")
UNDERSCORE_EXEMPT_COLUMNS = {"ID", "Parent Word ID", "URL", "Depth",
                             "Number in DED", "Number in CVOTGD"}


def fix_underscore_letters(text):
    return UNDERSCORE_RE.sub(lambda m: UNDERSCORE_LETTERS[m.group(1)], text)

# Starling dialect/source-split names -> canonical language in the
# three-tier tree. The original Starling name is preserved per row in the
# "Language (Starling)" extra-data field.
DIALECT_TO_LANGUAGE = {
    # Gondi dialects / sources
    "Koya Gondi": "Gondi",
    "Muria Gondi": "Gondi",
    "Maria Gondi": "Gondi",
    "Maria Gondi (Mitchell)": "Gondi",
    "Maria Gondi (Lind)": "Gondi",
    "Maria Gondi (Smith)": "Gondi",
    "Betul Gondi": "Gondi",
    "Adilabad Gondi": "Gondi",
    "Mandla Gondi": "Gondi",
    "Mandla Gondi (Phailbus)": "Gondi",
    "Mandla Gondi (Williamson)": "Gondi",
    "Seoni Gondi": "Gondi",
    "Gommu Gondi": "Gondi",
    "Yeotmal Gondi": "Gondi",
    "Chindwara Gondi": "Gondi",
    "Durg Gondi": "Gondi",
    "Chanda Gondi": "Gondi",
    # Kuwi sources (tree spells the language "Kuvi")
    "Kuwi (Schulze)": "Kuvi",
    "Kuwi (Fitzgerald)": "Kuvi",
    "Kuwi (Israel)": "Kuvi",
    "Kuwi (Mahanti)": "Kuvi",
    "Sunkarametta Kuwi": "Kuvi",
    "Parja Kuwi": "Kuvi",
    "Tekriya Kuwi": "Kuvi",
    "Dongriya Kuwi": "Kuvi",
    # Kui dialects
    "Khuttia Kui": "Kui",
    # Gadba varieties (tree distinguishes Ollari from Gadaba)
    "Ollari Gadba": "Ollari",
    "Salur Gadba": "Gadaba",
    "Kondekor Gadba": "Gadaba",
    "Poya Gadba": "Gadaba",
    # Kolami sources
    "Kinwat Kolami": "Kolami",
    "Kolami (Setumadhava Rao)": "Kolami",
    # Telugu varieties
    "Telugu (Krishnamurti)": "Telugu",
    "Inscriptional Telugu": "Telugu",
    "Merolu Telugu": "Telugu",
    # Konda sources
    "Konda (Burrow/Bhattacharya)": "Konda",
    # NOTE: Kasaba deliberately NOT mapped — Todd's call (2026-07-05): keep it
    # distinct on the first pass, erring toward granularity. It gets its own
    # language row via EXTRA_LANGUAGES below. Post-pilot: consider a way to
    # mark it as "dialect of Irula (uncertain)" on the entries themselves.
}

# Languages in the Starling data that aren't in the three-tier tree xlsx but
# should stay distinct (not merged into a tree language).
EXTRA_LANGUAGES = {
    "Kasaba": ("South", "Proto-South Dravidian I"),
}

# Tier labels shown as "Family: Subfamily" headers on the site. Todd asked to
# shorten these (2026-07-04): "Proto-South Dravidian: Proto-South Dravidian II"
# reads better as "South: Proto-South Dravidian II". Language names themselves
# are untouched.
FAMILY_LABELS = {
    "Proto-Dravidian": "Proto-Dravidian",
    "Proto-South Dravidian": "South",
    "Proto-Central Dravidian": "Central",
    "Proto-North Dravidian": "North",
}
SUBFAMILY_LABELS = {
    "Proto-South Dravidian I (South Dravidian)": "Proto-South Dravidian I",
    "Proto-South Dravidian II (South-Central Dravidian)": "Proto-South Dravidian II",
}
# Subfamilies that exist only to satisfy the three-tier schema (their family
# has no real subdivision). Named per the final sidebar review (2026-07-05) so
# headers read "Central: Central Dravidian" instead of "Central: Central".
SELF_SUBFAMILY_LABELS = {
    "Proto-South Dravidian": "Proto-South Dravidian",
    "Proto-Central Dravidian": "Central Dravidian",
    "Proto-North Dravidian": "North Dravidian",
    "Proto-Dravidian": "Proto-Dravidian",
}

# Intermediate proto-languages reconstructed in Starling but absent from the
# three-tier tree; placed under the tier they belong to (Krishnamurti 2003).
# Values are (family label, subfamily label) using the shortened names above.
PROTO_LANGUAGE_TIERS = {
    "Proto-South Dravidian": ("South", "Proto-South Dravidian"),
    "Proto-Central Dravidian": ("Central", "Central Dravidian"),
    "Proto-North Dravidian": ("North", "North Dravidian"),
    "Proto-Nilgiri": ("South", "Proto-South Dravidian I"),
    "Proto-Telugu": ("South", "Proto-South Dravidian II"),
    "Proto-Gondi-Kui": ("South", "Proto-South Dravidian II"),
    "Proto-Gondi": ("South", "Proto-South Dravidian II"),
    "Proto-Kui-Kuwi": ("South", "Proto-South Dravidian II"),
    "Proto-Pengo-Manda": ("South", "Proto-South Dravidian II"),
    "Proto-Kolami-Gadba": ("Central", "Central Dravidian"),
}

# The tree spreadsheet writes Tamil's subfamily without the "Proto-" prefix
# the other ten SD I rows use; normalize to the majority spelling.
SUBFAMILY_NORMALIZATION = {
    "South Dravidian I (South Dravidian)": "Proto-South Dravidian I (South Dravidian)",
}

# Records per batched upload file — the chunk size that reliably got the
# nahuatlex JSON through the admin Utilities uploader.
BATCH_SIZE = 1000

# The unpatched uploader on lrc-test throws on 'Etyma'/'HomographNumber'
# columns ("Etyma crosslinking not supported yet"). The compat export renames
# the link columns so they land harmlessly in extra data (displayed on word
# pages), and a later migration can turn them into real etyma links once the
# patched code is deployable.
COMPAT_KEY_RENAMES = {
    "IsEtymon": "Is Root",
    "HomographNumber": "Root Homograph",
    "Etyma": "Root Etymon",
    "EtymaHomographNumber": "Root Etymon Homograph",
}

# Optional review output of tag_buck_semantic_fields.py; when present, root
# rows get a 'Semantic Tag (Buck)' extra-data column.
BUCK_TAGS_CSV_NAME = "buck_tag_suggestions.csv"


def load_starling_rows():
    wb = openpyxl.load_workbook(STARLING_XLSX, read_only=True)
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    header = list(next(row_iter))
    rows = []
    dropped = 0
    kept_spaced_with_data = 0
    formless = []
    for values in row_iter:
        row = dict(zip(header, values))
        # Drop empty North Dravidian placeholders (see ARTIFACT_LANGUAGE) using
        # the RAW spelling, before normalization erases the hyphen distinction.
        if row.get("Language") == ARTIFACT_LANGUAGE:
            if is_nd_placeholder_artifact(row):
                dropped += 1
                continue
            # A spaced row that unexpectedly carries data is NOT a known
            # artifact — keep it and flag rather than silently discard.
            kept_spaced_with_data += 1
        for key in ("Language", "Parent Language"):
            if row.get(key) in LANGUAGE_NORMALIZATION:
                row[key] = LANGUAGE_NORMALIZATION[row[key]]
        for key, value in row.items():
            if isinstance(value, str) and "_" in value and key not in UNDERSCORE_EXEMPT_COLUMNS:
                row[key] = fix_underscore_letters(value)
        # Split source-qualified `form "gloss"` Headwords into a clean form plus a
        # preserved source gloss (word-379500 bug). Done after underscore repair
        # so both parts inherit it, and here at load time so the tree xlsx and the
        # import files stay consistent. The source gloss rides in a transient key
        # (not in `header`) so it never leaks into the tree xlsx or extra data.
        if row.get("Headword"):
            form, source_gloss = split_embedded_gloss(str(row["Headword"]).strip())
            row["Headword"] = form
            row["_meaning_source"] = source_gloss
            if source_gloss is None and '"' in form:
                formless.append(row.get("ID"))
        rows.append(row)
    wb.close()
    if dropped:
        print(f"dropped {dropped} empty Proto-North Dravidian placeholder rows")
    if kept_spaced_with_data:
        print(f"WARNING: kept {kept_spaced_with_data} spaced 'Proto-North Dravidian' "
              f"rows that carry data — review; not treated as artifacts")
    if formless:
        print(f"WARNING: {len(formless)} headwords are a bare quoted gloss with no "
              f"form — left unchanged, need the markup re-scrape: {formless}")
    return header, rows


def deduplicate_ids(rows):
    """Repair colliding row IDs from the destructuring notebook.

    'Proto-North Dravidian' and 'Proto-North-Dravidian' kept separate counters
    but shared the PND prefix, so 541 IDs collide across records, making
    Parent Word ID references ambiguous. The notebook always emits a parent
    row before its children, so each parent reference resolves to the most
    recent occurrence of that ID; colliding IDs get a '-2' suffix.
    """
    counts = {}
    latest = {}  # original ID -> unique ID of its most recent occurrence
    for row in rows:
        parent = row.get("Parent Word ID")
        if parent:
            row["Parent Word ID"] = latest.get(parent, parent)
        old = row["ID"]
        counts[old] = counts.get(old, 0) + 1
        new = old if counts[old] == 1 else f"{old}-{counts[old]}"
        latest[old] = new
        row["ID"] = new
    return rows


def propagate_ded_numbers(rows):
    """Give every protoform row the DED number(s) of its whole subtree."""
    children = defaultdict(list)
    by_id = {}
    for row in rows:
        by_id[row["ID"]] = row
        if row["Parent Word ID"]:
            children[row["Parent Word ID"]].append(row["ID"])

    def own_ded(row):
        value = row.get("Number in DED")
        if value is None or str(value).strip() == "":
            return set()
        return {int(part) for part in re.findall(r"\d+", str(value))}

    subtree_cache = {}

    def subtree_ded(node_id):
        if node_id in subtree_cache:
            return subtree_cache[node_id]
        numbers = own_ded(by_id[node_id])
        for child_id in children[node_id]:
            numbers |= subtree_ded(child_id)
        subtree_cache[node_id] = numbers
        return numbers

    for row in rows:
        if row["Language"].startswith("Proto-"):
            numbers = subtree_ded(row["ID"])
            row["Number in DED"] = ", ".join(str(n) for n in sorted(numbers))
        elif row.get("Number in DED") not in (None, ""):
            row["Number in DED"] = str(int(row["Number in DED"]))
    return rows


def write_batches(batch_dir, records):
    batch_dir.mkdir(exist_ok=True)
    for old in batch_dir.glob("dravidilex_batch_*.json"):
        old.unlink()
    n_batches = (len(records) + BATCH_SIZE - 1) // BATCH_SIZE
    for i in range(n_batches):
        chunk = records[i * BATCH_SIZE:(i + 1) * BATCH_SIZE]
        first, last = i * BATCH_SIZE + 1, i * BATCH_SIZE + len(chunk)
        name = f"dravidilex_batch_{i + 1:02d}_of_{n_batches}_entries_{first}-{last}.json"
        with open(batch_dir / name, "w", encoding="utf-8") as f:
            json.dump(chunk, f, ensure_ascii=False, indent=1)
    print(f"wrote {n_batches} chunks of <={BATCH_SIZE} to {batch_dir.relative_to(REPO_ROOT)}/")


def _xlsx_cell(value):
    """openpyxl can't write a list/dict cell — JSON-encode complex values
    (e.g. the Sources array) so the spreadsheet stays human-reviewable."""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def write_xlsx(path, header, rows):
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    ws.append(header)
    for row in rows:
        ws.append([_xlsx_cell(row.get(col)) for col in header])
    wb.save(path)


def build_languages_csv():
    """Family,Subfamily,Language rows for the Laravel import command.

    The tracked CSV is the practical source of truth across machines. The older
    xlsx is kept as a fallback for local historical workflows only.
    """
    entries = []
    seen = set()

    def add(family, subfamily, language):
        subfamily = subfamily or family
        subfamily = SUBFAMILY_NORMALIZATION.get(subfamily, subfamily)
        family = FAMILY_LABELS.get(family, family)
        subfamily = SELF_SUBFAMILY_LABELS.get(subfamily, SUBFAMILY_LABELS.get(subfamily, subfamily))
        key = (family, subfamily, language)
        if key not in seen:
            seen.add(key)
            entries.append(key)

    if LANGUAGES_CSV.exists():
        with open(LANGUAGES_CSV, encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                family = (row.get("Family") or "").strip()
                subfamily = (row.get("Subfamily") or "").strip()
                language = (row.get("Language") or "").strip()
                if family and language:
                    add(family, subfamily or None, language)
    elif TREE_XLSX.exists():
        wb = openpyxl.load_workbook(TREE_XLSX)
        ws = wb.active
        row_iter = ws.iter_rows(values_only=True)
        next(row_iter)  # header
        for family, subfamily, language in row_iter:
            if family is None:
                continue
            # Family- or subfamily-only tier rows carry no importable language;
            # proto-languages get real rows from PROTO_LANGUAGE_TIERS below.
            if language:
                add(family.strip(), subfamily.strip() if subfamily else None, language.strip())
        wb.close()
    else:
        raise FileNotFoundError(
            f"Missing language tiers: expected tracked CSV {LANGUAGES_CSV} "
            f"or fallback spreadsheet {TREE_XLSX}"
        )

    for language, (family, subfamily) in {**PROTO_LANGUAGE_TIERS, **EXTRA_LANGUAGES}.items():
        key = (family, subfamily, language)
        if key not in seen:
            seen.add(key)
            entries.append(key)

    return entries


def split_embedded_gloss(headword):
    """Separate a source-qualified reflex cell into (form, source gloss).

    StarlingDB's source-qualified rows (e.g. "Telugu (Krishnamurti)") render the
    form and a quoted gloss in one cell, which the destructuring notebook then
    stored whole as the Headword — e.g. `aḍalu "to be afraid, tremble, shake"`
    on word 379500. It affects a sixth of the pilot rows, concentrated in
    dialect-source languages (Gondi, Kuvi, Kolami, Kota, Parji).

    The form is everything before the first double quote; the gloss is what sits
    between the first and last quote (nested quotes inside a gloss are kept).
    The embedded gloss is usually MORE specific than the row's own Meaning (which
    carries the parent etymon's general sense), so it is worth preserving as a
    'Meaning (source)' extra-data field rather than discarding.

    Guard: a handful of rows are a bare quoted gloss with no form before it
    (e.g. `"hoe"`) — a distinct scrape defect that only the markup re-scrape can
    repair. Splitting those would blank the headword, so they are left untouched
    (form == original) and flagged by the caller.
    """
    if '"' not in headword:
        return headword, None
    form = headword.split('"', 1)[0].strip()
    if form == "":
        return headword, None  # form-less: leave unchanged, caller flags it
    first, last = headword.index('"'), headword.rindex('"')
    gloss = headword[first + 1:last] if last > first else headword[first + 1:]
    gloss = gloss.strip()
    return form, (gloss or None)


def etymon_entry(headword):
    """Etyma entries are stored without asterisks — including on inner
    variants like "*aḍái ~ *aḍí" — because the site's etymon views prepend a
    single <sup>*</sup> for the whole entry (IELex convention)."""
    return str(headword).replace("*", "").strip()


def headword_entries(headword):
    """Split reflex headwords only on top-level commas.

    Starling/DEDR headword cells often use parenthetical morphology such as
    `aṭai (-v-, -nt-)`. Those commas are display text, not entry separators, so
    the Laravel importer consumes this pre-shaped list instead of comma-splitting
    raw Headwords.
    """
    entries = []
    current = []
    depth = 0
    for char in str(headword):
        if char in "([":
            depth += 1
        elif char in ")]" and depth > 0:
            depth -= 1
        if char == "," and depth == 0:
            entry = "".join(current).strip()
            if entry:
                entries.append(entry)
            current = []
            continue
        current.append(char)
    entry = "".join(current).strip()
    if entry:
        entries.append(entry)
    return entries


def load_buck_tags():
    """Starling row ID -> (Buck field abbr, human-readable field label)."""
    path = OUT_DIR / BUCK_TAGS_CSV_NAME
    if not path.exists():
        return {}

    labels = {}
    with open(OUT_DIR / "buck_semantic_category.csv", encoding="utf-8-sig") as f:
        categories = {r["abbr"]: r["text"] for r in csv.DictReader(f)}
    with open(OUT_DIR / "buck_semantic_field.csv", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            abbr = r["abbr"]
            if abbr and not abbr.startswith("None"):
                category = categories.get(abbr.split("_")[0], "")
                labels[abbr] = f"{category} — {r['text']}" if category else r["text"]

    tags = {}
    with open(path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            abbr = row.get("chosen_abbr")
            if abbr:
                tags[row["Starling ID"]] = (abbr, labels.get(abbr, ""))
    return tags


class SourceHtmlSanitizer(HTMLParser):
    """Allow only source-entry markup that the lexicon word page should render."""

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.parts = []

    def handle_starttag(self, tag, attrs):
        if tag not in SOURCE_HTML_ALLOWED_TAGS:
            return
        if tag == "br":
            self.parts.append("<br>")
            return
        if tag == "a":
            href = ""
            for name, value in attrs:
                if name == "href" and value and value.startswith(("http://", "https://")):
                    href = value
                    break
            if href:
                self.parts.append(f'<a href="{html.escape(href, quote=True)}">')
            return
        self.parts.append(f"<{tag}>")

    def handle_endtag(self, tag):
        if tag in SOURCE_HTML_ALLOWED_TAGS and tag != "br":
            self.parts.append(f"</{tag}>")

    def handle_data(self, data):
        self.parts.append(html.escape(data))


def sanitize_source_html(source_html):
    sanitizer = SourceHtmlSanitizer()
    sanitizer.feed(source_html)
    sanitizer.close()
    return "".join(sanitizer.parts)


def clean_dedr_html(raw):
    """Normalize a scraped Burrow DEDR entry into clean, well-formed HTML.

    Keeps the linguistically meaningful `<b>`/`<i>` markup; strips the scrape's
    structural cruft: the `<div class='hw_result'>` wrappers, the `<number>`
    entry-number tag (covered by our source header + the reflex's `Number in
    DED`), the malformed `<xref="cdial">…</xref="cdial">` cross-reference
    pseudo-tags (text kept), and the trailing `<bibl>` line (our bold source
    header already carries the citation). The remaining HTML is allowlist
    sanitized before import because source text renders as HTML on word pages.
    """
    s = re.sub(r"</?div[^>]*>", "", raw)
    s = re.sub(r"<number>.*?</number>", "", s, flags=re.S)
    s = re.sub(r'</?xref="[^"]*">', "", s)
    s = re.sub(r"<bibl>.*?</bibl>", "", s, flags=re.S)
    s = re.sub(r"</?super>", lambda m: m.group(0).replace("super", "sup"), s)
    return sanitize_source_html(re.sub(r"\s+", " ", s).strip())


def load_dedr_entries():
    """DED number -> (page, cleaned_html) from the scraped Burrow corpus.

    Filtered to edition="DEDR" (the Appendix reuses the same numbers). A few
    DED numbers were scraped more than once; keep the fullest paragraph.
    """
    with open(BURROW_CORPUS, encoding="utf-8-sig") as f:
        corpus = json.load(f)
    best_len = {}
    index = {}
    for entry in corpus.get("entries", []):
        if entry.get("edition") != "DEDR":
            continue
        try:
            num = int(entry.get("ded_number"))
        except (TypeError, ValueError):
            continue
        raw = entry.get("raw_html") or ""
        if num in best_len and len(raw) <= best_len[num]:
            continue
        best_len[num] = len(raw)
        index[num] = (entry.get("page"), clean_dedr_html(raw))
    return index


def starling_url_metadata(url):
    query = parse_qs(urlparse(url).query)
    basename = query.get("basename", [""])[0]
    return STARLING_DATABASE_LABELS.get(basename, basename), query.get("text_number", [""])[0]


def build_starling_entry(record):
    """STARLING source body as a citation, not a duplicate entry.

    Word/gloss/tree details already render in the main table and Other Info.
    Keep only source-locator facts here: Starling branch/database, Starling ID,
    text number, and link to the original record.
    """
    parts = []
    url = record.get("URL")
    citation_bits = []
    if url:
        database_label, text_number = starling_url_metadata(url)
        if database_label:
            citation_bits.append(database_label)
        if text_number:
            citation_bits.append(f"text no. {text_number}")
    if record.get("Starling ID"):
        citation_bits.append(f"record {record['Starling ID']}")
    if citation_bits:
        parts.append(f"<small>{html.escape(' · '.join(citation_bits))}")
    if url:
        if citation_bits:
            parts.append("<br>")
        parts.append(f'<a href="{html.escape(url, quote=True)}">View original record</a>')
    if citation_bits:
        parts.append("</small>")
    return sanitize_source_html("".join(parts))


def build_sources(record, dedr_index, missing_ded):
    """The `Sources` array for one reflex: DEDR (full Burrow entry, per DED
    number), CVOTGD (page/entry number only — not scraped yet), and STARLING."""
    sources = []
    ded_value = record.get("Number in DED")
    if ded_value:
        for num in dict.fromkeys(int(x) for x in re.findall(r"\d+", str(ded_value))):
            entry = dedr_index.get(num)
            if entry is None:
                missing_ded[num] = missing_ded.get(num, 0) + 1
                continue
            page, cleaned = entry
            sources.append({
                "source": SOURCE_DEDR,
                "page_number": str(page) if page else "",
                "original_entry": cleaned,
            })
    cvotgd = record.get("Number in CVOTGD")
    if cvotgd:
        # Not scraped — a page/entry-number citation only (TODO: scrape CVOTGD).
        sources.append({
            "source": SOURCE_CVOTGD,
            "page_number": str(cvotgd).strip(),
            "original_entry": "",
        })
    if record.get("URL"):
        sources.append({
            "source": SOURCE_STARLING,
            "page_number": "",
            "original_entry": build_starling_entry(record),
        })
    return sources


def import_extra_key(column_name):
    """Normalize Starling/LRC extra-data key spelling for stable site display."""
    if column_name == "Additional Forms":
        return "Additional forms"
    if column_name == "ID":
        return "Starling ID"
    return column_name


def build_import_rows(header, rows, buck_tags, dedr_index):
    """Map tree rows onto the Utilities reflex-upload format.

    Parentless rows are marked `IsEtymon` (with a `HomographNumber`, since
    Starling reconstructs many identical roots); every descendant carries
    `Etyma` + `EtymaHomographNumber` pointing at its root, so the uploader can
    create the etymon→reflex tree. Roots always precede their descendants in
    file order, which batching into contiguous chunks preserves.

    Every reflex also gets a `Sources` array (DEDR full entry by DED number,
    CVOTGD number, STARLING); roots carry Starling only as an "Other Info" line,
    since etyma cannot hold sources.
    """
    extra_columns = [col for col in header if col not in ("Headword", "Meaning", "Language")]
    missing_ded = {}

    root_link = {}  # row ID -> (etymon entry, homograph number)
    homograph_counts = {}
    for row in rows:
        parent_id = row.get("Parent Word ID")
        if parent_id:
            root_link[row["ID"]] = root_link.get(parent_id)
        else:
            entry = etymon_entry(row["Headword"])
            homograph_counts[entry] = homograph_counts.get(entry, 0) + 1
            root_link[row["ID"]] = (entry, homograph_counts[entry])

    import_rows = []
    for row in rows:
        starling_language = row["Language"]
        language = DIALECT_TO_LANGUAGE.get(starling_language, starling_language)
        record = {
            "Headwords": str(row["Headword"]).strip(),
            "Gloss": str(row["Meaning"]).strip() if row.get("Meaning") else "",
            "Language": language,
        }
        if language != starling_language:
            record["Language (Starling)"] = starling_language
        source_gloss = row.get("_meaning_source")
        if source_gloss and source_gloss.strip().lower() != record["Gloss"].strip().lower():
            record["Meaning (source)"] = source_gloss
        link = root_link.get(row["ID"])
        if not row.get("Parent Word ID"):
            record["IsEtymon"] = "1"
            record["EtymonEntry"] = link[0]
            record["HomographNumber"] = str(link[1])
            tag = buck_tags.get(row["ID"])
            if tag:
                record["Semantic Tag (Buck)"] = tag[0]
            # Etyma cannot hold sources; surface Starling provenance as Other Info.
            record["Source (StarlingDB)"] = ROOT_STARLING_ATTRIBUTION
        elif link:
            record["HeadwordEntries"] = headword_entries(record["Headwords"])
            record["Etyma"] = link[0]
            record["EtymaHomographNumber"] = str(link[1])
        for col in extra_columns:
            if not row.get("Parent Word ID") and col == "Depth":
                continue
            value = row.get(col)
            if value is None or str(value).strip() == "":
                continue
            key = import_extra_key(col)
            text_value = str(value).strip()
            if key in record and record[key] != text_value:
                record[key] = f"{record[key]}\n{text_value}"
            else:
                record[key] = text_value
        if row.get("Parent Word ID"):
            sources = build_sources(record, dedr_index, missing_ded)
            if sources:
                record["Sources"] = sources
        import_rows.append(record)
    if missing_ded:
        print(f"WARNING: {sum(missing_ded.values())} reflex-DED references point at "
              f"{len(missing_ded)} DED numbers absent from the DEDR corpus "
              f"(no DEDR source attached): {sorted(missing_ded)[:15]}")
    return import_rows


def main():
    OUT_DIR.mkdir(exist_ok=True)

    header, rows = load_starling_rows()
    rows = deduplicate_ids(rows)
    rows = propagate_ded_numbers(rows)

    tree_path = OUT_DIR / "dravidian_starling_data.xlsx"
    write_xlsx(tree_path, header, rows)
    print(f"wrote {tree_path.relative_to(REPO_ROOT)} ({len(rows)} rows)")

    languages = build_languages_csv()
    lang_path = OUT_DIR / "dravidilex_languages.csv"
    with open(lang_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Family", "Subfamily", "Language"])
        writer.writerows(languages)
    print(f"wrote {lang_path.relative_to(REPO_ROOT)} ({len(languages)} languages)")

    buck_tags = load_buck_tags()
    if buck_tags:
        print(f"injecting {len(buck_tags)} Buck semantic tags onto root rows")
    dedr_index = load_dedr_entries()
    print(f"loaded {len(dedr_index)} DEDR entries for the Sources column")
    import_rows = build_import_rows(header, rows, buck_tags, dedr_index)

    known_languages = {language for _, _, language in languages}
    # Etyma/root rows are not imported as LexReflex records, so their Language
    # value does not need a LexLanguage row. In particular, the built-in
    # protolanguage page already represents Proto-Dravidian; adding it here as a
    # normal language duplicates it in the sidebar.
    unmapped = sorted(
        {r["Language"] for r in import_rows if not r.get("IsEtymon")} - known_languages
    )
    if unmapped:
        raise SystemExit(f"languages missing from languages CSV: {unmapped}")

    json_path = OUT_DIR / "dravidilex_batch_import.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(import_rows, f, ensure_ascii=False, indent=1)
    print(f"wrote {json_path.relative_to(REPO_ROOT)} ({len(import_rows)} records)")

    # Chunked copies for the admin Utilities uploader (the nahuatlex-proven
    # pattern) — upload in ascending order so roots precede their reflexes.
    write_batches(OUT_DIR / "batched", import_rows)

    # Compat variant for the UNPATCHED uploader on lrc-test: link columns are
    # renamed so they land in extra data instead of triggering the
    # "Etyma crosslinking not supported yet" exception. Flat import — no
    # etyma pages — but the tree stays visible/searchable per word.
    compat_rows = [
        {COMPAT_KEY_RENAMES.get(key, key): value for key, value in record.items()}
        for record in import_rows
    ]
    write_batches(OUT_DIR / "batched_compat_lrctest", compat_rows)

    import_columns = ["Headwords", "Gloss", "Language", "Language (Starling)"]
    for record in import_rows:
        for key in record:
            if key not in import_columns:
                import_columns.append(key)
    xlsx_path = OUT_DIR / "dravidilex_batch_import.xlsx"
    write_xlsx(xlsx_path, import_columns, import_rows)
    print(f"wrote {xlsx_path.relative_to(REPO_ROOT)} ({len(import_rows)} rows)")


if __name__ == "__main__":
    main()
