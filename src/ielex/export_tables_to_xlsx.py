#!/usr/bin/env python3
"""Export every analysis table from the IELEX notes + notebooks into one
multi-tab Excel workbook for sharing in Google Drive (auto-converts to a
Google Sheet with one tab per table).

Sources:
  - Obsidian notes:  <VAULT>/IELEX/*.md
      * GitHub-flavored markdown (pipe) tables
      * ASCII box-drawing tables inside ``` code fences (the Discord-formatted ones)
  - Notebooks:       src/ielex/notebooks/dpp_analysis.ipynb,
                     src/ielex/notebooks/dpp_followup_analysis.ipynb
      * markdown tables in markdown cells
      * rendered DataFrame outputs in code cells

Output: <VAULT>/IELEX/_exports/IELEX_analysis_tables.xlsx
First sheet is an "Index" listing every table and its source.

Run with the project venv (needs pandas, openpyxl, lxml, bs4):
    lrc_env/bin/python src/ielex/export_tables_to_xlsx.py
"""
from __future__ import annotations

import io
import json
import re
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd

# --------------------------------------------------------------------------- paths
VAULT = Path("/Users/kevinbarcenas/Downloads/Obsidian/LinguisticResearchCenter")
NOTES_DIR = VAULT / "IELEX"
NB_DIR = Path(__file__).resolve().parent / "notebooks"
OUT_DIR = NOTES_DIR / "_exports"
OUT_FILE = OUT_DIR / "IELEX_DPP_apples_source_of_truth.xlsx"
MD_FILE = OUT_DIR / "IELEX_chat_tables.md"  # GFM headline set for pasting into the chat
LEGACY_ARCHIVE = OUT_DIR / "IELEX_analysis_tables.xlsx"  # the old 46-tab dump, removed

# Headline set for the chat paste: (tab shortname, heading, one-line caption). No em dashes.
MD_HEADLINE = [
    ("Coverage results", "Coverage - where the gap actually is",
     "Roots are ~complete; the real gap is reflexes (681 empty roots, 112 fillable now)."),
    ("Apples-to-apples gap", "Apples-to-apples (Pokorny-content only, 1,357 shared roots)",
     "Two reflex sets are similar in size but overlap only ~25-30%, and miss in opposite directions."),
    ("Reflex sources", "Which book the LRC cited (this is the 1,100 'Pokorny' figure)",
     "Citation bookkeeping, NOT content: only 1,100 reflexes have 'Pokorny' typed in the source box."),
    ("Reflex sources 2", "Where IELex out-counts DPP, the surplus is Webster's + AHD",
     "The IELex-only forms are modern-English derivatives, not richer Pokorny."),
    ("Languages absent", "Top languages missing from matched roots",
     "The gap is structural: whole language columns absent, not stray forms."),
]

AS_OF = "2026-06-07"  # latest 'updated' date across the source notes/notebooks
MAX_SHEET = 31         # Excel hard limit on worksheet-name length

THEME1 = "1. DPP analysis - coverage & gap"
THEME2 = "2. Apples-to-apples crosswalk"

# Curated source-of-truth set: only the FINAL analysis tables, grouped into the two
# themes the team needs. Superseded directional figures (the Q-B 54% / ~14,700 count
# tables and the count-based 'top under-populated roots') are deliberately excluded;
# the apples-to-apples content comparison (section B) replaces them.
# Each rule = (source_substr, key_substr, theme, short_tab_name).
# key matches the table's section OR title; short_tab_name is the worksheet label.
CURATED = [
    # ---- Theme 1: DPP analysis - coverage & the real (reflex) gap ----------
    ("Follow-up Q&A", "Q-A.", THEME1, "Ballpark"),
    ("DPP Crosswalk", "Coverage", THEME1, "Coverage results"),
    ("dpp_analysis", "Gap Analysis", THEME1, "Empty-root gap"),
    ("dpp_analysis", "Top 20 languages by reflex", THEME1, "DPP top languages"),
    ("Q-B Reflex Gap", "absent most often", THEME1, "Languages absent"),
    ("DPP Crosswalk", "Mapping draft", THEME1, "Abbreviation map"),
    # ---- Theme 2: Apples-to-apples - where the crosswalk stands ------------
    ("Apples-to-Apples", "Pokornys", THEME2, "The 5 Pokornys"),
    ("Reply to colleague", "A. The reflexes where we beat", THEME2, "Reflex sources"),
    ("Reply to colleague", "B. What this does to the gap", THEME2, "Apples-to-apples gap"),
    ("Reply to colleague", "C. This makes Todd's tiering", THEME2, "Tiering plan"),
    ("Reply to colleague", "F. AHD (Watkins)", THEME2, "AHD handling"),
    ("dpp_followup", "tiering plan", THEME2, "Tiering (notebook)"),
    ("Apples-to-Apples", "Known holes", THEME2, "Known holes"),
]

# Hand-written headline summary (the FINAL numbers) for the front tab. No em dashes.
KEY_NUMBERS = [
    ("ROOTS - essentially complete", "", ""),
    ("IELex roots (etyma)", "2,222", "~ Pokorny book total (~1,183 pp x ~2/pg). 'Missing ~20%' is NOT roots."),
    ("DPP unique roots", "1,518", "Fewer than IELex; DPP is not a superset."),
    ("IELex roots with 0 reflexes", "681 (30.6%)", "The real gap is reflexes, not headwords."),
    ("...immediately fillable from DPP", "112", "After homograph-aware join; 569 need page-join / stricter match."),
    ("", "", ""),
    ("APPLES-TO-APPLES - Pokorny-content only, 1,357 shared roots", "", ""),
    ("Distinct reflex forms in BOTH (overlap)", "~11,900", "Only ~25-30% overlap."),
    ("IELex-only forms", "~40,000", "62.7% Webster's 7th + 27.0% AHD: the modern-English 'psychology' layer."),
    ("Pokorny-only forms (we lack)", "~38,500", "Older daughter-lang forms: Anglo-Saxon, Old Icelandic, NHG, Lith."),
    ("IELex reflexes that cite Pokorny", "1.7% (1,100)", "Citation is not content: bookkeeping, not Pokorny coverage."),
    ("IELex reflexes actually Pokorny-attested", "~28% (15,867 / 56,271)", "IELex = multi-source lexicon on Pokorny's root skeleton, not a reproduction."),
    ("", "", ""),
    ("DIRECTION & PLAN", "", ""),
    ("Gap points both ways", "merge", "Pull in ~38.5k Pokorny-only forms; demote Webster's/AHD English to Tier 2 (source-tagged)."),
    ("Laryngeals (h1h2h3)", "not from DPP/Starling", "All classic 1959 Pokorny. Sources: Declan's corrections + AHD (which has them)."),
    ("Provenance", "single digitization", "DPP, PIET, indogermanisch.org all = same Starostin/Lubotsky StarlingDB digitization, not independent."),
    ("", "", ""),
    ("CONFIDENCE", "directional, not audit-grade", "Matcher ~8% false-negative floor; vs partial DPP (1,517 roots, ~10% blank-root rows dropped). See 'Known holes' tab."),
    ("As of", AS_OF, "Latest update across source notes/notebooks. Superseded Q-B 54% / ~14,700 count figures intentionally excluded."),
]


@dataclass
class Table:
    title: str          # human-readable table title
    source: str         # note / notebook filename
    section: str        # nearest heading
    df: pd.DataFrame
    sheet: str = ""     # assigned later
    theme: str = ""     # curated theme grouping
    shortname: str = "" # desired short worksheet label


# --------------------------------------------------------------------------- cleanup
_WIKILINK = re.compile(r"\[\[([^\]|]+)\|([^\]]+)\]\]")          # [[a|b]] -> b
_WIKILINK2 = re.compile(r"\[\[([^\]]+)\]\]")                    # [[a]]   -> a
_MDLINK = re.compile(r"\[([^\]]+)\]\([^)]+\)")                 # [t](u)  -> t
_HTMLTAG = re.compile(r"</?[a-zA-Z][^>]*>")                     # <b> etc.


def clean_cell(s: str) -> str:
    """Strip markdown/HTML noise so cells read clean in Sheets; keep unicode."""
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\\|", "|")
    s = s.replace("—", "-").replace("–", "-")  # no em/en dashes
    s = _WIKILINK.sub(r"\2", s)
    s = _WIKILINK2.sub(r"\1", s)
    s = _MDLINK.sub(r"\1", s)
    s = _HTMLTAG.sub("", s)
    s = s.replace("**", "").replace("`", "")
    return s.strip()


# --------------------------------------------------------------------------- titling
_HEADING = re.compile(r"^\s*#{1,6}\s+(.*?)\s*#*\s*$")
_BOLD_ONLY = re.compile(r"^\s*\*\*(.+?)\*\*.*$")  # a line that starts with a bold label


def _title_from_context(lines: list[str], idx: int, heading: str) -> str:
    """Look just above a table for a bold label (e.g. **lex_etyma**) used to
    disambiguate the several tables that share one heading; else use the heading."""
    for k in range(idx - 1, max(idx - 4, -1), -1):
        ln = lines[k].strip()
        if not ln:
            continue
        m = _BOLD_ONLY.match(ln)
        if m:
            return clean_cell(m.group(1))
        break  # only inspect the immediately preceding non-blank line
    return heading


# --------------------------------------------------------------------------- GFM tables
def parse_markdown_tables(text: str, source: str, default_heading: str = "(top)"):
    """Yield Table objects for every GFM pipe table in *text*."""
    lines = text.splitlines()
    heading = default_heading
    out: list[Table] = []
    i = 0
    while i < len(lines):
        ln = lines[i]
        hm = _HEADING.match(ln)
        if hm:
            heading = clean_cell(hm.group(1))
        is_row = ln.lstrip().startswith("|")
        is_sep = i + 1 < len(lines) and re.match(r"^\s*\|?[\s:|-]+\|[\s:|-]*$", lines[i + 1])
        if is_row and is_sep:
            start = i
            header = _split_pipe(lines[i])
            i += 2  # skip header + separator
            rows = []
            while i < len(lines) and lines[i].lstrip().startswith("|"):
                rows.append(_split_pipe(lines[i]))
                i += 1
            df = _rows_to_df(header, rows)
            if df is not None:
                title = _title_from_context(lines, start, heading)
                out.append(Table(title=title, source=source, section=heading, df=df))
            continue
        i += 1
    return out


def _split_pipe(line: str) -> list[str]:
    s = line.strip()
    if s.startswith("|"):
        s = s[1:]
    if s.endswith("|"):
        s = s[:-1]
    return [clean_cell(c) for c in s.split("|")]


def _rows_to_df(header, rows):
    if not header:
        return None
    width = len(header)
    norm = []
    for r in rows:
        if len(r) < width:
            r = r + [""] * (width - len(r))
        elif len(r) > width:
            r = r[:width]
        norm.append(r)
    if not norm:
        return None
    return pd.DataFrame(norm, columns=_dedupe_headers(header))


def _dedupe_headers(header):
    seen, out = {}, []
    for h in header:
        h = h or "col"
        if h in seen:
            seen[h] += 1
            out.append(f"{h}.{seen[h]}")
        else:
            seen[h] = 0
            out.append(h)
    return out


# --------------------------------------------------------------------------- ASCII box tables
_BOX_BORDER_CHARS = set("╔╦╗╠╬╣╚╩╝═╤╧╪┌┬┐├┼┤└┴┘─╒╕╘╛")
_VBAR = "║"  # the column delimiter used in these notes


def _is_border(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    return all(ch in _BOX_BORDER_CHARS or ch.isspace() for ch in stripped)


def _box_cells(line: str) -> list[str]:
    s = line.strip()
    s = s.strip(_VBAR)
    return [clean_cell(c) for c in s.split(_VBAR)]


def parse_box_tables(text: str, source: str):
    """Yield Table objects for ASCII box-drawing tables found in ``` fences."""
    lines = text.splitlines()
    heading = "(top)"
    out: list[Table] = []
    in_fence = False
    fence_lines: list[str] = []
    fence_heading = heading
    seen_under_heading: dict[str, int] = {}

    def flush(buf, head):
        # A fence may hold one box table.
        rows = [l for l in buf if _VBAR in l and not _is_border(l)]
        if len(rows) < 2:
            return
        matrix = [_box_cells(l) for l in rows]
        width = max(len(r) for r in matrix)
        matrix = [r + [""] * (width - len(r)) for r in matrix]
        header = matrix[0]
        body = matrix[1:]
        merged: list[list[str]] = []
        for r in body:
            # continuation line if any cell is blank (wrapped text / spanning row)
            is_cont = merged and any(c == "" for c in r)
            if is_cont:
                for j, c in enumerate(r):
                    if c:
                        merged[-1][j] = (merged[-1][j] + " " + c).strip()
            else:
                merged.append(list(r))
        df = _rows_to_df(header, merged)
        if df is None:
            return
        n = seen_under_heading.get(head, 0) + 1
        seen_under_heading[head] = n
        title = head if n == 1 else f"{head} ({n})"
        out.append(Table(title=clean_cell(title), source=source, section=head, df=df))

    for ln in lines:
        hm = _HEADING.match(ln)
        if hm and not in_fence:
            heading = clean_cell(hm.group(1))
        if ln.strip().startswith("```"):
            if not in_fence:
                in_fence = True
                fence_lines = []
                fence_heading = heading
            else:
                in_fence = False
                flush(fence_lines, fence_heading)
            continue
        if in_fence:
            fence_lines.append(ln)
    return out


# --------------------------------------------------------------------------- notebooks
def _drop_serialized_index(df: pd.DataFrame) -> pd.DataFrame:
    """read_html turns a DataFrame's RangeIndex into a leading 'Unnamed: 0'
    column of 0,1,2,…  Drop it; keep it when it holds real row labels."""
    if df.shape[1] == 0:
        return df
    first = df.columns[0]
    if isinstance(first, str) and first.startswith("Unnamed"):
        col = pd.to_numeric(df[first], errors="coerce")
        if col.notna().all() and list(col.astype(int)) == list(range(len(df))):
            return df.drop(columns=[first])
    return df


def parse_notebook(path: Path):
    nb = json.loads(path.read_text())
    source = path.name
    out: list[Table] = []
    heading = "(top)"
    for cell in nb.get("cells", []):
        src = "".join(cell.get("source", []))
        if cell.get("cell_type") == "markdown":
            for hm in _HEADING.finditer(src):
                heading = clean_cell(hm.group(1))
            out.extend(parse_markdown_tables(src, source, default_heading=heading))
        elif cell.get("cell_type") == "code":
            first_comment = ""
            for line in src.splitlines():
                if line.strip().startswith("#"):
                    first_comment = clean_cell(line.strip("# ").strip())
                    break
            for o in cell.get("outputs", []):
                html = o.get("data", {}).get("text/html")
                if not html:
                    continue
                html = "".join(html)
                if "</table>" not in html.lower():
                    continue
                try:
                    dfs = pd.read_html(io.StringIO(html))
                except ValueError:
                    continue
                for df in dfs:
                    df = _drop_serialized_index(df)
                    df = df.astype(object).where(pd.notnull(df), "")
                    title = first_comment or heading or "DataFrame"
                    out.append(Table(title=title, source=source, section=heading, df=df))
    return out


# --------------------------------------------------------------------------- sheet names
_BAD = re.compile(r"[\[\]:*?/\\]")


def assign_sheet_names(tables: list[Table]):
    used = {"Index", "Key numbers"}
    for t in tables:
        base = _BAD.sub(" ", t.shortname or t.title).strip() or "Table"
        base = re.sub(r"\s+", " ", base)[:MAX_SHEET].strip()
        name = base
        n = 1
        while name.lower() in {u.lower() for u in used}:
            n += 1
            suffix = f" {n}"
            name = base[: MAX_SHEET - len(suffix)].strip() + suffix
        assert len(name) <= MAX_SHEET, f"sheet name too long: {name!r}"
        used.add(name)
        t.sheet = name


# --------------------------------------------------------------------------- curate
def curate(tables: list[Table]) -> list[Table]:
    """Select only the final source-of-truth tables, in CURATED order, tagging theme."""
    picked: list[Table] = []
    used = set()
    for src_sub, key_sub, theme, short in CURATED:
        for i, t in enumerate(tables):
            if i in used:
                continue
            if src_sub.lower() in t.source.lower() and (
                key_sub.lower() in t.section.lower() or key_sub.lower() in t.title.lower()
            ):
                t.theme = theme
                t.shortname = short
                picked.append(t)
                used.add(i)
    return picked


# --------------------------------------------------------------------------- markdown
def _gfm(df: pd.DataFrame, index: bool = False) -> str:
    def esc(v):
        return ("" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)).replace("|", "\\|").replace("\n", " ").strip()

    cols = ([str(df.index.name or "")] if index else []) + [str(c) for c in df.columns]
    out = ["| " + " | ".join(esc(c) for c in cols) + " |",
           "| " + " | ".join("---" for _ in cols) + " |"]
    for idx, row in df.iterrows():
        cells = ([str(idx)] if index else []) + [esc(v) for v in row]
        out.append("| " + " | ".join(cells) + " |")
    return "\n".join(out)


def write_markdown(tables: list[Table], out_file: Path):
    by_sheet = {t.sheet: t for t in tables}
    parts = [
        "# IELEX - DPP analysis & apples-to-apples (headline tables)",
        f"_Source of truth as of {AS_OF}. Full set in the Drive spreadsheet._",
        "",
        "## Key numbers",
        "",
    ]
    # Key numbers as a clean GFM table (skip blank separator rows; bold section rows).
    rows = []
    for fnd, num, note in KEY_NUMBERS:
        if not fnd:
            continue
        if not num and not note:
            rows.append((f"**{fnd}**", "", ""))
        else:
            rows.append((fnd, num, note))
    key_df = pd.DataFrame(rows, columns=["Finding", "Number", "Note"])
    parts.append(_gfm(key_df))
    parts.append("")
    for sheet, heading, caption in MD_HEADLINE:
        t = by_sheet.get(sheet)
        if t is None:
            continue
        parts += [f"## {heading}", f"_{caption}_", "", _gfm(t.df, index=not isinstance(t.df.index, pd.RangeIndex)), ""]
    out_file.write_text("\n".join(parts))


# --------------------------------------------------------------------------- write
def write_workbook(tables: list[Table], out_file: Path):
    out_file.parent.mkdir(parents=True, exist_ok=True)
    key = pd.DataFrame(KEY_NUMBERS, columns=["Finding", "Number", "Note"])
    def _nd(s):  # display label with no em/en dashes
        return s.replace("—", "-").replace("–", "-")

    index = pd.DataFrame(
        [[_nd(t.theme), t.sheet, _nd(t.title), _nd(t.source)] for t in tables],
        columns=["Theme", "Sheet", "Table title", "Source"],
    )
    with pd.ExcelWriter(out_file, engine="openpyxl") as xw:
        key.to_excel(xw, sheet_name="Key numbers", index=False)
        index.to_excel(xw, sheet_name="Index", index=False)
        for t in tables:
            keep_index = not isinstance(t.df.index, pd.RangeIndex)
            t.df.to_excel(xw, sheet_name=t.sheet, index=keep_index)
        _format(xw)
    return index


def _format(xw):
    from openpyxl.utils import get_column_letter

    for ws in xw.book.worksheets:
        ws.freeze_panes = "A2"
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            longest = 0
            for cell in ws[letter]:
                if cell.value is not None:
                    longest = max(longest, max((len(p) for p in str(cell.value).split("\n")), default=0))
            ws.column_dimensions[letter].width = min(max(longest + 2, 10), 70)


# --------------------------------------------------------------------------- main
def main():
    tables: list[Table] = []

    for md in sorted(NOTES_DIR.glob("*.md")):
        text = md.read_text()
        tables.extend(parse_markdown_tables(text, md.name))
        tables.extend(parse_box_tables(text, md.name))

    for nb in ["dpp_analysis.ipynb", "dpp_followup_analysis.ipynb"]:
        p = NB_DIR / nb
        if p.exists():
            tables.extend(parse_notebook(p))

    # drop empty tables
    tables = [t for t in tables if not t.df.empty and t.df.shape[1] > 0]

    # keep only the curated source-of-truth set, grouped by theme
    curated = curate(tables)
    assign_sheet_names(curated)
    index = write_workbook(curated, OUT_FILE)
    write_markdown(curated, MD_FILE)

    # remove the old noisy 46-tab archive so there's a single source of truth
    if LEGACY_ARCHIVE.exists():
        LEGACY_ARCHIVE.unlink()
        print(f"Removed legacy archive → {LEGACY_ARCHIVE.name}")

    print(f"Wrote {len(curated)} tables (+ Key numbers + Index) → {OUT_FILE}")
    for theme, grp in index.groupby("Theme", sort=False):
        print(f"\n{theme}")
        for _, r in grp.iterrows():
            print(f"   - {r['Table title'][:60]}")
    longest = max(len(t.sheet) for t in curated)
    print(f"\nLongest tab name: {longest} chars (Excel max {MAX_SHEET}).")


if __name__ == "__main__":
    main()
